# -*- coding: utf-8 -*-
from openpyxl import load_workbook

from gerpar import parecer_contrato_prorrogacao_servicos_conntinuados

def make_parecer(data):
    contrato = data[0]
    processo = data[1]
    contratada = data[2]
    cnpj = data[3]
    objeto = data[4]
    prazo_prorrogacao = data[5]
    data = data[6]

    # 1 Serviços continuados
    # 2 Locação
    # 3 Obras
    # 4 Termos de Credenciamento
    parecer_contrato_prorrogacao_servicos_conntinuados(1, contrato, processo, contratada, cnpj, objeto, prazo_prorrogacao, data)


def main():
    wb = load_workbook(filename='relacao_contratos.xlsx')
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        make_parecer(row)

if __name__ == "__main__":
   main()