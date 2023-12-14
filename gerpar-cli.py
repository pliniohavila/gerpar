# -*- coding: utf-8 -*-
from docx import Document
from openpyxl import load_workbook

from gerpar import parecer_edital, parecer_contrato

def make_parecer(data):
    numero_contrato = data[0]
    contratado = data[2]
    cpf_cnpj = data[3]
    objeto = data[4]
    prazo_prorrogacao = data[5]
    data_pedido_aditivo = data[6]

    # São gerados os pareceres
    parecer_edital(requerente, assunto, processo, num_modalidade, dataEdital)
    parecer_contrato(requerente, assunto, processo, num_modalidade, dataContrato)


def main():
    wb = load_workbook(filename='relacao_contratos.xlsx')
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        make_parecer(row)

if __name__ == "__main__":
   main()